# excel_utils.py
import os
from openpyxl import Workbook, load_workbook

EXCEL_FILE = "company_data.xlsx"
HEADERS = [
    "Company Name", "Name", "Designation", "Website", "Email", "City", "State",
    "Region", "Address", "Pincode", "Telephone", "Manufacturing Product/Services"
]

def save_to_excel(tab_line: str):
    if not tab_line.strip(): return  # skip empty results

    row_data = tab_line.strip().split("\t")

    # Create new workbook if file doesn't exist
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Companies"
        ws.append(HEADERS)
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Companies"]

    ws.append(row_data)
    wb.save(EXCEL_FILE)
    print("✅ Saved to Excel.")
